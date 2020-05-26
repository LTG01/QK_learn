import openpyxl

# 创建一个工作蒲
workbook = openpyxl.load_workbook('示例.xlsx')
# 打印所有的表明
print(workbook.sheetnames)

sheet = workbook[workbook.sheetnames[-1]]
print(sheet.max_column)
print(sheet.max_row)
# 取第一行所有的值
# excel 是从1开始计算
for i in range(1, sheet.max_row+1):
    # row
    print(sheet.cell(row=i, column=1).value)

workbook.save('实例（修改）.xlsx')