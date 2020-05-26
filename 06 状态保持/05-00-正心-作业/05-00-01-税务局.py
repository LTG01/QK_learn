"""
    目标地址：http://www.chinatax.gov.cn/chinatax/n810346/n810825/index.html
    1. 采集
    采集以下栏目：
        增值税、消费税的前两页数据
    每个栏目的：
        标题、发文日期、文号

    2. 保存
        将所有信息保存到`税务局.xlsx`文件，依据栏目名创建数据表，每个栏目的数据保存到对应的数据表
"""
import time
import openpyxl
import requests

url = 'http://www.chinatax.gov.cn/api/query'
params = {
    "siteCode": "bm29000fgk",
    "tab": "all",
    "key": "9A9C42392D397C5CA6C1BF07E2E0AA6F"
}


def get_data(cate, page):
    return {
        "timeOption": "0",
        "C6": cate,  # 消费税 "增值税"
        "page": str(page),  # 1 2 3
        "pageSize": "10",
        "keyPlace": "1",
        "sort": "dateDesc",
        "qt": "*",
    }


headers = {
    "Cookie": "_Jo0OQK=2A9770CF074FBBD9D79D285DE9AD8356007732AA709673F2B59C19EA3FCA7585AF315A682DB07A992DCFB4007BB4BFF805A097C03810E1A055CE736E577D5D7B30101C83797AB67C7627EF4EAE0C6A2C5647EF4EAE0C6A2C56491A0E8A83004FA3FGJ1Z1cA==; yfx_c_g_u_id_10003701=_ck20052119111112901515278360493; CPS_SESSION=4347F6AB31474ABB597A8BC3E60FB3AD; yfx_f_l_v_t_10003701=f_t_1590059471285__r_t_1590059471285__v_t_1590061325937__r_c_0",
    "Host": "www.chinatax.gov.cn",
    "Origin": "http://www.chinatax.gov.cn",
    "Referer": "http://www.chinatax.gov.cn/chinatax/n810346/n810825/index.html",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36",
}


def save_data(result, sheet1):
    result_list = result['resultList']
    for r in result_list:

        sheet1.append([r['publishTime'], r['dreTitle']])


wb = openpyxl.Workbook()

sheet1 = wb.create_sheet('消费税')
sheet2 = wb.create_sheet('增值税')
for page in range(1, 4):
    time.sleep(3)
    data = get_data(cate='消费税', page=page)
    response = requests.post(url=url, params=params, data=data, headers=headers)
    result = response.json()
    save_data(result, sheet1)

    data1 = get_data(cate='增值税', page=page)
    response = requests.post(url=url, params=params, data=data1, headers=headers)
    result = response.json()
    save_data(result, sheet2)

    wb.save('税务局.xlsx')
