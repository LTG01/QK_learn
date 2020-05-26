import openpyxl

# 创建一个工作蒲
workbook = openpyxl.Workbook()

# 创建一张表
sheet1 = workbook.create_sheet('表1')
sheet2 = workbook.create_sheet('九九乘法表')

# sheet1["A1"] = "A1"
# sheet1["B1"] = "B1"
# cell 单元格对象 row col value
sheet1.cell(row=2, column=1).value = "A1"
sheet1.cell(row=2, column=2).value = "B1"
sheet1.append([1, 2, 3, 4, 5])

for i in range(1, 10):
    for j in range(1, i + 1):
        print(f"{j} * {i} = {i * j}", end="\t")
        sheet2.cell(row=j, column=i).value = f"{j} * {i} = {i * j}"
    print()

# 保存 如果名字与源文件是一样的，就会覆盖
workbook.save('示例.xlsx')
